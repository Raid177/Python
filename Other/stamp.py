import fitz
import subprocess
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont

# === Вхідні дані ===
source_path = r"C:\Users\la\OneDrive\Рабочий стол\На оплату!\test.txt"
target_folder = r"C:\Users\la\OneDrive\Рабочий стол\На оплату!\Оплачено"

# === Визначення типу ===
ext = os.path.splitext(source_path)[1].lower()

# === Відкриваємо відповідний редактор ===
if ext == ".pdf":
    print("🔄 Відкриваємо PDF у редакторі...")
    viewer = subprocess.Popen([
        r"C:\Program Files\Tracker Software\PDF Editor\PDFXEdit.exe",
        source_path
    ])
elif ext in (".xls", ".xlsx"):
    print("🔄 Відкриваємо Excel у редакторі...")
    viewer = subprocess.Popen([
        r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
        source_path
    ])
elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
    print("🔄 Відкриваємо зображення...")
    viewer = subprocess.Popen([
        r"C:\Program Files (x86)\FastStone Image Viewer\FSViewer.exe",
        source_path
    ])
elif ext == ".txt":
    print("🔄 Відкриваємо текстовий файл...")
    viewer = subprocess.Popen([
        "notepad.exe",
        source_path
    ])
else:
    print("❌ Непідтримуваний тип файлу.")
    exit()

viewer.wait()
print("✅ Редактор закрито.")

# === Підтвердження ===
root = tk.Tk()
root.withdraw()
root.attributes("-topmost", True)
answer = messagebox.askyesno("Підтвердження", "Внести штамп 'PAID'?")

if not answer:
    print("❌ Внесення штампа скасовано.")
    exit()

# === Створення цільової папки ===
os.makedirs(target_folder, exist_ok=True)
ts = datetime.now().strftime("%Y-%m-%d %H-%M")
filename = os.path.basename(source_path)
base_name, original_ext = os.path.splitext(filename)
new_name = f"{ts} {base_name}{ext if ext != '.xls' else '.xlsx'}"
new_path = os.path.join(target_folder, new_name)

if ext == ".pdf":
    print("✍️ Вносимо штамп у PDF...")
    doc = fitz.open(source_path)
    page = doc[0]

    stamp_text = "PAID"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    x = page.rect.width - 200
    y = 50
    fontsize = 18
    color = (0, 0.6, 0)

    page.insert_text((x, y), stamp_text, fontsize=fontsize, color=color)
    page.insert_text((x, y + 25), timestamp, fontsize=fontsize, color=color)

    doc.save(new_path, garbage=4, deflate=True)
    doc.close()
    os.remove(source_path)
    print(f"✅ Штамп додано і файл переміщено до:\n{new_path}")

elif ext == ".xls":
    print("ℹ Конвертуємо .xls → .xlsx")
    try:
        df = pd.read_excel(source_path, engine="xlrd")
        df.to_excel(new_path, index=False)
        print("✅ Конвертація успішна")
    except Exception as e:
        print(f"❌ Не вдалося конвертувати .xls: {e}")
        exit()

elif ext == ".xlsx":
    shutil.copy2(source_path, new_path)

if ext in (".xls", ".xlsx"):
    print("✍️ Вносимо штамп у Excel...")
    wb = load_workbook(new_path)
    sheet = wb.active
    stamp = f"PAID {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sheet.insert_rows(1)
    sheet["A1"] = stamp
    sheet["A1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    wb.save(new_path)
    wb.close()
    os.remove(source_path)
    print(f"✅ Штамп додано і Excel-файл переміщено до:\n{new_path}")

elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
    print("✍️ Вносимо штамп у зображення...")
    image = Image.open(source_path).convert("RGB")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype("arialbd.ttf", 20)
    stamp_text = "PAID"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    x, y = image.width - 220, 20
    draw.text((x, y), stamp_text, font=font, fill=(0, 153, 0))
    draw.text((x, y + 25), timestamp, font=font, fill=(0, 153, 0))
    image.save(new_path)
    os.remove(source_path)
    print(f"✅ Штамп додано і зображення переміщено до:\n{new_path}")

elif ext == ".txt":
    print("✍️ Вносимо штамп у текстовий файл...")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(source_path, "r", encoding="utf-8") as f:
        original_content = f.read()
    with open(new_path, "w", encoding="utf-8") as f:
        f.write(f"PAID {timestamp}\n\n{original_content}")
    os.remove(source_path)
    print(f"✅ Штамп додано і текстовий файл переміщено до:\n{new_path}")
