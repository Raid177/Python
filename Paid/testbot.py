import fitz
import subprocess
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont
import requests
import pymysql
from dotenv import load_dotenv
import msvcrt
import sys
import tempfile

# === Блокування запуску другого екземпляра ===
lockfile_path = os.path.join(tempfile.gettempdir(), "stamp_paid.lock")
try:
    lock_file = open(lockfile_path, 'w')
    msvcrt.locking(lock_file.fileno(), msvcrt.LK_NBLCK, 1)
except OSError:
    print("⚠️ Програма вже запущена. Другий екземпляр не дозволено.")
    sys.exit()

# === Завантаження конфігурації з .env ===
load_dotenv("C:/Users/la/OneDrive/Pet Wealth/Analytics/Python_script/.env")

BOT_TOKEN = os.getenv("BOT_TOKEN")
API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
DB_HOST = os.getenv("DB_HOST")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_DATABASE = os.getenv("DB_DATABASE")
FALLBACK_CHAT_ID = os.getenv("FALLBACK_CHAT_ID", "-4624165634")

log_file = os.path.join(os.path.dirname(__file__), "stamp_log.txt")

# === Підключення до БД ===
conn = pymysql.connect(
    host=DB_HOST,
    user=DB_USER,
    password=DB_PASSWORD,
    database=DB_DATABASE,
    charset='utf8mb4',
    autocommit=True
)
cursor = conn.cursor()

def log(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{timestamp}] {message}"
    print(line)
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def send_telegram_reply(file_name, new_name):
    try:
        log(f"🔍 Пошук у БД за file_name: {file_name}")
        cursor.execute("""
            SELECT chat_id, message_id FROM telegram_files
            WHERE file_name = %s
            ORDER BY id DESC LIMIT 1
        """, (file_name,))
        row = cursor.fetchone()
        if row:
            chat_id, message_id = row
            log(f"📦 Знайдено: chat_id={chat_id}, message_id={message_id}")
            msg = f"✅ Файл «{file_name}» оплачено"

            # Перевірка доступності чату
            test_resp = requests.get(
                f"https://api.telegram.org/bot{BOT_TOKEN}/getChat",
                params={"chat_id": chat_id}
            )
            log(f"🛠️ getChat response: {test_resp.status_code} {test_resp.text}")

            # Пробуємо відповісти на повідомлення
            resp = requests.post(API_URL, data={
                "chat_id": chat_id,
                "text": msg,
                "reply_to_message_id": message_id
            })

            if resp.status_code == 400:
                log("⚠️ Неможливо відповісти на повідомлення. Надсилаю окремо.")
                resp = requests.post(API_URL, data={
                    "chat_id": chat_id,
                    "text": msg
                })

            log(f"📤 Telegram response: {resp.status_code} {resp.text}")

            cursor.execute("""
                UPDATE telegram_files SET status='paid' WHERE file_name = %s
            """, (file_name,))
        else:
            msg = (
                f"✅ Файл «{file_name}» оплачено\n"
                f"⚠️ Але файл не знайдено в базі.\n‼️ Надсилайте файли через бота, щоб бот міг відповісти на оригінал."
            )
            resp = requests.post(API_URL, data={
                "chat_id": FALLBACK_CHAT_ID,
                "text": msg
            })
            log(f"📤 Telegram (fallback) response: {resp.status_code} {resp.text}")
    except Exception as e:
        log(f"❌ Не вдалося надіслати повідомлення в Telegram: {e}")


# === Основна логіка взаємодії з файлами ===
last_dir = os.getcwd()

while True:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_paths = filedialog.askopenfilenames(
        title="Оберіть файли для обробки",
        initialdir=last_dir,
        filetypes=[
            ("Всі підтримувані", "*.pdf *.xls *.xlsx *.jpg *.jpeg *.png *.bmp *.tiff *.txt"),
            ("PDF файли", "*.pdf"),
            ("Excel файли", "*.xls *.xlsx"),
            ("Зображення", "*.jpg *.jpeg *.png *.bmp *.tiff"),
            ("Текстові файли", "*.txt")
        ]
    )

    if not file_paths:
        print("👋 Завершення роботи: файли не обрано.")
        break

    last_dir = os.path.dirname(file_paths[0])

    for i, source_path in enumerate(file_paths, 1):
        target_folder = os.path.join(os.path.dirname(source_path), "Оплачено")
        ext = os.path.splitext(source_path)[1].lower()

        try:
            if ext == ".pdf":
                viewer = subprocess.Popen([r"C:\\Program Files\\Tracker Software\\PDF Editor\\PDFXEdit.exe", source_path])
            elif ext in (".xls", ".xlsx"):
                viewer = subprocess.Popen([r"C:\\Program Files\\Microsoft Office\\root\\Office16\\EXCEL.EXE", source_path])
            elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
                viewer = subprocess.Popen([r"C:\\Program Files (x86)\\FastStone Image Viewer\\FSViewer.exe", source_path])
            elif ext == ".txt":
                viewer = subprocess.Popen(["notepad.exe", source_path])
            else:
                log(f"IGNORED: Непідтримуваний файл {source_path}")
                continue

            viewer.wait()

            filename = os.path.basename(source_path)
            answer = messagebox.askyesno("Підтвердження", f"Обробити файл «{filename}» як оплачений 'PAID'?\nФайл {i} із {len(file_paths)}")
            if not answer:
                log(f"SKIPPED: {filename} — без змін")
                continue

            os.makedirs(target_folder, exist_ok=True)
            ts = datetime.now().strftime("%Y-%m-%d %H-%M")
            base_name, original_ext = os.path.splitext(filename)
            new_name = f"{ts} {base_name}{ext if ext != '.xls' else '.xlsx'}"
            new_path = os.path.join(target_folder, new_name)

            if ext == ".pdf":
                doc = fitz.open(source_path)
                page = doc[0]
                stamp_text = "PAID"
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                x = page.rect.width - 200
                y = 50
                fontsize = 18
                color = (0, 0.6, 0)
                page.insert_text((x, y), stamp_text, fontsize=fontsize, color=color)
                page.insert_text((x, y + 25), timestamp, fontsize=fontsize, color=color)
                doc.save(new_path, garbage=4, deflate=True)
                doc.close()
                os.remove(source_path)
            elif ext == ".xls":
                df = pd.read_excel(source_path, engine="xlrd")
                df.to_excel(new_path, index=False)
            elif ext == ".xlsx":
                shutil.copy2(source_path, new_path)

            if ext in (".xls", ".xlsx"):
                wb = load_workbook(new_path)
                sheet = wb.active
                stamp = f"PAID {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                sheet.insert_rows(1)
                sheet["A1"] = stamp
                sheet["A1"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                wb.save(new_path)
                wb.close()
                os.remove(source_path)
            elif ext in (".jpg", ".jpeg", ".png", ".bmp", ".tiff"):
                image = Image.open(source_path).convert("RGB")
                draw = ImageDraw.Draw(image)
                font = ImageFont.truetype("arialbd.ttf", 20)
                stamp_text = "PAID"
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                x, y = image.width - 220, 20
                draw.text((x, y), stamp_text, font=font, fill=(0, 153, 0))
                draw.text((x, y + 25), timestamp, font=font, fill=(0, 153, 0))
                image.save(new_path)
                os.remove(source_path)
            elif ext == ".txt":
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
                with open(source_path, "r", encoding="utf-8") as f:
                    original_content = f.read()
                with open(new_path, "w", encoding="utf-8") as f:
                    f.write(f"PAID {timestamp}\n\n{original_content}")
                os.remove(source_path)

            log(f"✔ OPLACHENO | {filename} → {new_name}")
            send_telegram_reply(filename, new_name)

        except Exception as err:
            log(f"❌ ERROR обробки {os.path.basename(source_path)}: {err}")
