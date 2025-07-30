# file_utils/office.py
"""
Модуль для вставки штампа "Paid ДатаЧас" у документи Excel (XLSX, XLS).
"""

import os
from datetime import datetime
from log import log

import openpyxl
from openpyxl.styles import PatternFill, Font

import xlrd
import xlwt
import xlutils.copy

def stamp_excel(file_path: str, stamp: str):
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Додаємо штамп у A1
            ws.insert_rows(1)
            ws["A1"] = stamp.replace("Оплачено", "Paid")

            # Стиль: синій фон, великий шрифт
            ws["A1"].fill = PatternFill(start_color="D9EAFD", end_color="D9EAFD", fill_type="solid")
            ws["A1"].font = Font(size=20, bold=True, color="000000")

            wb.save(file_path)
            log(f"🟦 XLSX: Додано штамп у {file_path}")
        except Exception as e:
            log(f"❌ XLSX помилка: {e}")

    elif ext == ".xls":
        try:
            rb = xlrd.open_workbook(file_path, formatting_info=True)
            wb = xlutils.copy.copy(rb)
            ws = wb.get_sheet(0)

            # Вставляємо штамп у A1 (рядок 0, колонка 0)
            style = xlwt.XFStyle()

            # Стиль: синій фон, чорний шрифт
            font = xlwt.Font()
            font.name = 'Arial'
            font.bold = True
            font.height = 20 * 20  # 20pt
            style.font = font

            pattern = xlwt.Pattern()
            pattern.pattern = xlwt.Pattern.SOLID_PATTERN
            pattern.pattern_fore_colour = 44  # Light turquoise
            style.pattern = pattern

            ws.write(0, 0, stamp.replace("Оплачено", "Paid"), style)

            wb.save(file_path)
            log(f"🟩 XLS: Додано штамп у {file_path}")
        except Exception as e:
            log(f"❌ XLS помилка: {e}")

    else:
        log(f"⚠️ Непідтримуваний формат Excel: {file_path}")


# 🔧 Тест
if __name__ == "__main__":
    test_xlsx = r"C:\Users\la\OneDrive\Рабочий стол\test.xlsx"
    test_xls = r"C:\Users\la\OneDrive\Рабочий стол\test.xls"
    now = datetime.now().strftime("Оплачено %Y-%m-%d %H:%M")
    stamp_excel(test_xlsx, now)
    stamp_excel(test_xls, now)
